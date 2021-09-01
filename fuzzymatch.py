# This program is free software licensed under the terms of GNU GPL v2
# Author: Pavel Urusov

import argparse
import enum
import shutil
import sys

from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

COLOR_LITERAL_MATCH = "90EE90"
COLOR_FUZZY_HIGH_SCORE = "FCE883"
COLOR_FUZZY_LOW_SCORE = "FF91A4"


class MessageType(enum.Enum):
    GENERAL = 0,
    INFO = 1,
    ERROR = 2


def fancy_message(message: str, type: MessageType) -> str:
    prefix = ""
    if type == MessageType.INFO:
        prefix = "[i] "
    if type == MessageType.ERROR:
        prefix = "[!] "
    if type == MessageType.GENERAL:
        prefix = "[*] "
    return prefix + message


def new_file_name(fname: str, suffix: str) -> str:
    last_dot = fname.rfind('.')
    if last_dot == -1:
        return fname + "_" + suffix
    return fname[:last_dot] + "_" + suffix + fname[last_dot:]


def find_match(string_to_match: str, d: dict, w: bool) -> tuple:
    max_score = 0
    matched_string = ""
    for potential_match in d.keys():
        score = fuzz.WRatio(string_to_match,
                            potential_match) if w else fuzz.QRatio(
                                string_to_match, potential_match)
        if score >= max_score:
            max_score = score
            matched_string = potential_match
    return (matched_string, max_score)


def build_source_dict(sheet: Worksheet, match_column: str, source_column: str,
                      min_row: int, max_row: int) -> dict:
    source_dict = {}
    row_count = min_row
    match_index = column_index_from_string(match_column) - 1
    source_index = column_index_from_string(source_column) - 1
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row):
        sys.stdout.write(
            fancy_message(f"Source document: reading row {row_count}\r",
                          MessageType.GENERAL))
        sys.stdout.flush()
        key = str(row[match_index].value).strip()
        value = row[source_index].value
        source_dict[key] = value
        row_count += 1
    return source_dict


def update_dest(sheet: Worksheet, match_column: str, dest_column: str,
                min_row: int, max_row: int, source: dict, score_threshold: int,
                weighted: bool):
    for i in range(min_row, max_row + 1):
        sys.stdout.write(
            fancy_message(f"Destination document: updating row {i}\r",
                          MessageType.GENERAL))
        sys.stdout.flush()
        cell_match = match_column + str(i)
        cell_dest = dest_column + str(i)
        string_to_match = str(sheet[cell_match].value).strip()

        # try literal match at first
        if string_to_match != "" and string_to_match in source.keys():
            # only update cells that do not match the source already
            if sheet[cell_dest].value != source[string_to_match]:
                sheet[cell_dest] = source[string_to_match]
                color_fill = PatternFill(fill_type="solid",
                                         start_color=COLOR_LITERAL_MATCH)
                sheet[cell_dest].fill = color_fill
        else:  # now try fuzzy matching
            matched_string, score = find_match(string_to_match, source,
                                               weighted)
            if score >= score_threshold:
                # only update cells that do not match the source already
                if sheet[cell_dest].value != source[matched_string]:
                    sheet[cell_dest] = source[matched_string]
                    # if ratio is 99 or 100, set the background color to yellow, otherwise set it to red
                    fill_color = COLOR_FUZZY_LOW_SCORE if score < 99 else COLOR_FUZZY_HIGH_SCORE
                    color_fill = PatternFill(fill_type="solid",
                                             start_color=fill_color)
                    sheet[cell_dest].fill = color_fill


def get_workbook(path: str, ro: bool) -> Workbook:
    try:
        wb = load_workbook(path, read_only=ro)
        return wb
    except:
        message = fancy_message(f"There was an error opening file {path}",
                                MessageType.ERROR)
        print(message)
        sys.exit(2)


def save_workbook(wb: Workbook, path: str):
    try:
        wb.save(path)
    except:
        message = fancy_message(f"There was an error saving file {path}",
                                MessageType.ERROR)
        print(message)
        sys.exit(2)


def get_args() -> argparse.Namespace:
    argparser = argparse.ArgumentParser(
        description=
        "Match two Excel documents by contents using fuzzy string matching.")
    argparser.add_argument("dest", help="destination document")
    argparser.add_argument("source", help="source document")
    argparser.add_argument("-o",
                           "--output",
                           type=str,
                           nargs="?",
                           const="",
                           default="None",
                           help="output document")
    argparser.add_argument(
        "--dest-match",
        type=str,
        nargs="?",
        const="B",
        default="B",
        help=
        "column in the destination document used to match the content (default: B)"
    )
    argparser.add_argument(
        "--source-match",
        type=str,
        nargs="?",
        const="W",
        default="W",
        help=
        "column in the source document used to match the content (default: W)")
    argparser.add_argument(
        "--dest-column",
        type=str,
        nargs="?",
        const="G",
        default="G",
        help=
        "column in the destination document which will be populated (default: G)"
    )
    argparser.add_argument(
        "--source-column",
        type=str,
        nargs="?",
        const="AE",
        default="AE",
        help=
        "column in the source document used as the source of data (default: AE)"
    )
    argparser.add_argument(
        "--dest-min-row",
        type=int,
        nargs="?",
        const="2",
        default="2",
        help="min row in the destination document (default: 2)")
    argparser.add_argument("--source-min-row",
                           type=int,
                           nargs="?",
                           const="2",
                           default="2",
                           help="min row in the source document (default: 2)")
    argparser.add_argument(
        "--dest-max-row",
        type=int,
        nargs="?",
        const="-1",
        default="-1",
        help="max row in the destination document (default: actual max row)")
    argparser.add_argument(
        "--source-max-row",
        type=int,
        nargs="?",
        const="-1",
        default="-1",
        help="max row in the source document (default: actual max row)")
    argparser.add_argument("-n",
                           "--no-backup",
                           help="do not backup the destination document",
                           action='store_true')
    argparser.add_argument(
        "-t",
        "--threshold",
        type=int,
        nargs="?",
        const=90,
        default=90,
        help="minimum score that will be considered a match (default: 90)")
    argparser.add_argument(
        "-w",
        "--weighted",
        help=
        "use weighted ratio instead of simple ratio for calculating scores",
        action='store_true')
    return argparser.parse_args()


def parse_arguments(args: argparse.Namespace) -> tuple:
    # set the output file name
    if args.output == "":
        output_file_name = new_file_name(args.dest, "new")
    elif args.output.lower() == "none":
        output_file_name = args.dest
    else:
        output_file_name = args.output

    # set destination file name and source file name
    dest_file_name = args.dest
    source_file_name = args.source

    # set columns
    dest_match_column = args.dest_match.upper()
    dest_column = args.dest_column.upper()
    source_match_column = args.source_match.upper()
    source_column = args.source_column.upper()

    # set min rows
    dest_min_row = args.dest_min_row
    source_min_row = args.dest_min_row

    # set max rows
    source_max_row = args.source_max_row
    dest_max_row = args.dest_max_row

    # other settings
    threshold = args.threshold
    no_backup = args.no_backup
    weighted = args.weighted

    return (dest_file_name, source_file_name, output_file_name,
            dest_match_column, dest_column, source_match_column, source_column,
            dest_min_row, dest_max_row, source_min_row, source_max_row,
            threshold, no_backup, weighted)


def main():
    #get the command line arguments and parse them
    args = get_args()
    dest_file_name, source_file_name, output_file_name, dest_match_column, dest_column, source_match_column, source_column, dest_min_row, dest_max_row, source_min_row, source_max_row, threshold, no_backup, weighted = parse_arguments(
        args)

    # if the output file is the same as the destination file and the --no-backup flag is not set, create a backup copy of the destination file
    if output_file_name == dest_file_name and not no_backup:
        backup_name = new_file_name(dest_file_name, "old")
        shutil.copyfile(dest_file_name, backup_name)

    # open the files and get active worksheets
    dest = get_workbook(dest_file_name, False)
    source = get_workbook(source_file_name, True)
    dest_sheet = dest.active
    source_sheet = source.active

    # inform the user about requested options
    s = f"Minimum score that will be considered a match: {threshold}."
    print(fancy_message(s, MessageType.INFO))
    s = "Using weighted ratio to calculate scores." if weighted else "Using simple ratio to calculate scores."
    print(fancy_message(s, MessageType.INFO))

    # if max rows not set by user, get the actual max rows
    if source_max_row == -1:
        source_max_row = source_sheet.max_row
    print(
        fancy_message(
            f"Source document: using rows {source_min_row} to {source_max_row}",
            MessageType.INFO))
    if dest_max_row == -1:
        dest_max_row = dest_sheet.max_row
    print(
        fancy_message(
            f"Destination document: using rows {dest_min_row} to {dest_max_row}",
            MessageType.INFO))

    # this dictionary holds the data from the source file
    source_dict = build_source_dict(source_sheet, source_match_column,
                                    source_column, source_min_row,
                                    source_max_row)
    print(
        fancy_message("Source document: all rows read successfully",
                      MessageType.INFO))

    # update the destination worksheet
    update_dest(dest_sheet, dest_match_column, dest_column, dest_min_row,
                dest_max_row, source_dict, threshold, weighted)
    print(
        fancy_message("Destination document: all rows updated successfully",
                      MessageType.INFO))

    # save the destination file
    print(fancy_message(f"Saving file: {output_file_name}", MessageType.INFO))
    save_workbook(dest, output_file_name)


if __name__ == "__main__":
    main()
